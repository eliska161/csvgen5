from flask import Flask, request, send_file, render_template
import pandas as pd
from openpyxl import load_workbook
import os
from tempfile import NamedTemporaryFile

app = Flask(__name__)

# Definerer filplasseringen for regnearkmalen
EXCEL_TEMPLATE_PATH = os.path.join('assets', 'Regnskapsark-for-elevbedrifter.xlsx')

@app.route('/')
def upload_page():
    return render_template('index.html')  # Sender brukeren til opplastingsskjemaet

@app.route('/process', methods=['POST'])
def process_files():
    # Sjekk om CSV-filen er lastet opp
    csv_file = request.files['csv_file']

    if not csv_file:
        return "Feil: CSV-fil må lastes opp", 400

    # Les inn CSV-dataene
    csv_data = pd.read_csv(csv_file)
    csv_data['Dato'] = pd.to_datetime(csv_data['Dato']).dt.strftime('%d.%m.%Y')  # Konverterer til norsk datoformat
    csv_data['Ut'] = csv_data['Beløp'].apply(lambda x: -x if x < 0 else 0)
    csv_data['Inn'] = csv_data['Beløp'].apply(lambda x: x if x > 0 else 0)

    # Åpne Excel-malen
    wb = load_workbook(EXCEL_TEMPLATE_PATH)
    ws = wb['Ark1']  # Velg arket der dataene skal settes inn

    # Sett data i de riktige cellene i Excel
    for i, row in csv_data.iterrows():
        if 6 + i > 19:
            break  # Stopper hvis vi overskrider antall rader som skal fylles
        ws[f"A{6 + i}"] = f"A{6 + i}"  # Fyller inn Billag (automatisert i henhold til rekkefølgen)
        ws[f"B{6 + i}"] = row['Dato']
        ws[f"C{6 + i}"] = row['Beskrivelse']
        ws[f"D{6 + i}"] = row['Ut']
        ws[f"E{6 + i}"] = row['Inn']

    # Lagre det oppdaterte regnearket til en midlertidig fil
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        temp_path = temp_file.name
        wb.save(temp_path)

    # Send filen tilbake til brukeren for nedlasting
    return send_file(temp_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=8080)
